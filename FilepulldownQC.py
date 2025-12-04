# -*- coding: utf-8 -*-
"""
FilepulldownQC (FED / BEAM)

FED mode:
- Folder picker
- Scans FED###_*.csv files
- Verifies 4 files per FED and presence of Bandit100, Bandit80, FR1, PR1
- Flags header-only (empty) CSVs (~1.5 KB or 0 data rows)
- Reports unique device counts (overall and per task)
- Moves files into task folders: Bandit_100, Bandit_80, FR1, PR1
- Uses key's FED3 column to check whether each FED is represented in each task:
    * For each FED ID (from files and from key) report:
        - In_Key, In_Files
        - Has_Bandit100, Has_Bandit80, Has_FR1, Has_PR1
- Outputs fed_session_check_report.xlsx with red-highlighted "bad" cells.
- Creates ZIPs for each task folder (if non-empty) named: Gene_Gene_ID_Task_L0.zip

BEAM mode (robust + date range filter):
- Single ZIP picker
- Asks for a date via calendar (GUI):
    * If that date is a Saturday, keeps Saturday + Sunday (two days).
    * Otherwise keeps that one date.
- Extracts to a temp folder
- Recursively finds ALL .csv anywhere inside (no folder-name assumptions)
- If none, auto-extracts nested ZIPs into sibling temp folders and searches again
- Moves ONLY CSVs whose filenames match: BEAMXXX_YYYYMMDD??.csv and start<=YYYYMMDD<=end
- Places matched CSVs into sibling folder "BEAM" (same layer as the original zip)
- After completion, deletes all temporary extract / nested folders so ONLY:
    * the BEAM folder
    * the Gene_GeneID_BEAM_L0.zip
  remain as outputs (plus the original zip)
- Creates a ZIP from the BEAM folder named Gene_GeneID_BEAM_L0.zip if a key is loaded
- Creates beam_session_check_report.xlsx comparing BEAM IDs in files vs BEAM column in key,
  with red-highlighted "bad" cells.
"""

import os
import re
import sys
import shutil
import zipfile
import subprocess
from datetime import timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

APP_VERSION = "1.2.1"

# For the uploaded key (Excel)
# beam_ids: set of 3-digit strings from BEAM column in key (if present)
# fed_ids:  set of 3-digit strings from FED3 column in key (if present)
KEY_INFO = {
    "gene": None,
    "gene_id": None,
    "path": None,
    "beam_ids": set(),
    "fed_ids": set(),
}


class UserCancel(Exception):
    """Raised when the user cancels a file/date selection in the GUI."""
    pass


# =========================
# Shared helpers
# =========================

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)
    return path

def unique_dest_path(dest_dir, filename):
    base, ext = os.path.splitext(filename)
    candidate = os.path.join(dest_dir, filename)
    i = 1
    while os.path.exists(candidate):
        candidate = os.path.join(dest_dir, f"{base} ({i}){ext}")
        i += 1
    return candidate

def is_numeric_dirname(name):
    return bool(re.fullmatch(r"\d+", name))

def open_folder_in_explorer(path):
    """Open the given folder in the system file explorer."""
    try:
        path = os.path.abspath(path)
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception as e:
        print(f"Could not open folder in file explorer: {e}")


# ---------- Excel styling helpers ----------

def style_fed_report_excel(path):
    """
    Open fed_session_check_report.xlsx and highlight:
      - Any non-empty 'Missing_Types' cell
      - Any False/false cell in:
          Count_OK(=4), All_Types_OK, Overall_OK, In_Key, In_Files,
          Has_Bandit100, Has_Bandit80, Has_FR1, Has_PR1
    """
    if not os.path.isfile(path):
        return
    wb = load_workbook(path)
    ws = wb.active

    header = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        header[str(cell.value)] = col_idx

    red_fill = PatternFill(
        start_color="FFFF9999",
        end_color="FFFF9999",
        fill_type="solid",
    )

    bool_cols = [
        "Count_OK(=4)",
        "All_Types_OK",
        "Overall_OK",
        "In_Key",
        "In_Files",
        "Has_Bandit100",
        "Has_Bandit80",
        "Has_FR1",
        "Has_PR1",
    ]
    missing_col = "Missing_Types"

    for row_idx in range(2, ws.max_row + 1):
        # Missing_Types: highlight if non-empty
        if missing_col in header:
            c = ws.cell(row=row_idx, column=header[missing_col])
            val = c.value
            if val not in (None, ""):
                c.fill = red_fill

        # Boolean-ish columns: highlight False
        for col_name in bool_cols:
            if col_name not in header:
                continue
            c = ws.cell(row=row_idx, column=header[col_name])
            val = c.value
            if isinstance(val, bool):
                if not val:
                    c.fill = red_fill
            elif isinstance(val, str):
                if val.strip().lower() == "false":
                    c.fill = red_fill

    wb.save(path)


def style_beam_report_excel(path):
    """
    Open beam_session_check_report.xlsx and highlight any False/false cells
    in columns In_Key and In_Files.
    """
    if not os.path.isfile(path):
        return
    wb = load_workbook(path)
    ws = wb.active

    header = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        header[str(cell.value)] = col_idx

    red_fill = PatternFill(
        start_color="FFFF9999",
        end_color="FFFF9999",
        fill_type="solid",
    )

    bool_cols = ["In_Key", "In_Files"]

    for row_idx in range(2, ws.max_row + 1):
        for col_name in bool_cols:
            if col_name not in header:
                continue
            c = ws.cell(row=row_idx, column=header[col_name])
            val = c.value
            if isinstance(val, bool):
                if not val:
                    c.fill = red_fill
            elif isinstance(val, str):
                if val.strip().lower() == "false":
                    c.fill = red_fill

    wb.save(path)


# =========================
# FED implementation
# =========================

REQUIRED_TYPES = {"Bandit100", "FR1", "Bandit80", "PR1"}
TASK_DIRNAME = {
    "Bandit100": "Bandit_100",
    "Bandit80": "Bandit_80",
    "FR1": "FR1",
    "PR1": "PR1",
}
FED_PATTERN = re.compile(r"(?i)\bFED(\d{3})_", re.ASCII)
EMPTY_SIZE_BYTES = 1500  # ~1KB header-only files


def extract_fed_digits(filename):
    m = FED_PATTERN.search(filename)
    return m.group(1) if m else None


def normalize_session_type(value):
    if value is None:
        return None
    v = str(value).strip()
    for canonical in REQUIRED_TYPES:
        if v.lower() == canonical.lower():
            return canonical
    return None


def is_likely_empty(csv_path):
    try:
        size_ok = os.path.getsize(csv_path) <= EMPTY_SIZE_BYTES
    except OSError:
        size_ok = False

    empty_by_rows = False
    try:
        head = pd.read_csv(csv_path, dtype=str, nrows=5, encoding_errors="ignore")
        if head.shape[0] == 0:
            empty_by_rows = True
        else:
            nonempty = (
                head.apply(lambda s: s.astype(str).str.strip())
                .replace({"": pd.NA})
                .dropna(how="all")
            )
            empty_by_rows = nonempty.empty
    except Exception:
        empty_by_rows = False

    return size_ok, empty_by_rows


def read_session_type(csv_path):
    try:
        head = pd.read_csv(csv_path, dtype=str, nrows=1000, encoding_errors="ignore")
        cols_map = {c.lower(): c for c in head.columns}
        if "session_type" not in cols_map:
            df = pd.read_csv(csv_path, dtype=str, encoding_errors="ignore")
            cols_map = {c.lower(): c for c in df.columns}
            if "session_type" not in cols_map:
                return None
            series = df[cols_map["session_type"]]
        else:
            series = head[cols_map["session_type"]]

        series = series.dropna().astype(str).str.strip()
        norm = series.map(normalize_session_type).dropna()
        if norm.empty:
            return None
        return norm.mode().iloc[0]
    except Exception:
        return None


# ---- console versions of these three helpers (used in CLI mode) ----

def pick_folder():
    path = input("Enter FED folder path: ").strip()
    if not path or not os.path.isdir(path):
        print("No valid folder selected. Exiting.")
        sys.exit(0)
    return path


def pick_zip_file():
    path = input("Enter path to a .zip file: ").strip()
    if not (path and os.path.isfile(path) and path.lower().endswith(".zip")):
        print("Invalid .zip path. Exiting.")
        sys.exit(1)
    return path


def ask_date_range_to_keep():
    pattern = re.compile(r"^\d{8}$")
    while True:
        s = input("Enter START date (YYYYMMDD): ").strip()
        e = input("Enter END date (YYYYMMDD): ").strip()
        if not (pattern.fullmatch(s) and pattern.fullmatch(e)):
            print("Dates must be exactly 8 digits (YYYYMMDD). Try again.")
            continue
        if s > e:
            print("START must be <= END. Try again.")
            continue
        return s, e


def move_file_to_task_folder(folder, filename, session_type):
    canonical = normalize_session_type(session_type)
    if canonical not in TASK_DIRNAME:
        return False
    dest_dir = ensure_dir(os.path.join(folder, TASK_DIRNAME[canonical]))
    src_path = os.path.join(folder, filename)
    dest_path = unique_dest_path(dest_dir, filename)
    try:
        shutil.move(src_path, dest_path)
        return True
    except Exception as e:
        print(f"  ! Could not move {filename} → {dest_dir}: {e}")
        return False


# ---- core FED logic (takes a folder, no GUI) ----

def run_fed_pipeline_core(folder):
    all_files = [f for f in os.listdir(folder) if f.lower().endswith(".csv")]
    csv_files = [f for f in all_files if FED_PATTERN.search(f)]
    if not csv_files:
        print("No CSV files with pattern 'FED###_' found in the selected folder.")
        return

    per_fed = {}
    all_types_global = {}
    file_to_type = {}
    file_is_empty = set()

    for fname in csv_files:
        fed = extract_fed_digits(fname)
        if not fed:
            continue
        fullpath = os.path.join(folder, fname)
        empty_by_size, empty_by_rows = is_likely_empty(fullpath)
        stype = None if (empty_by_size or empty_by_rows) else read_session_type(fullpath)

        fed_info = per_fed.setdefault(
            fed, {"files": [], "types": set(), "by_type": {}, "empties": []}
        )
        fed_info["files"].append(fname)

        if empty_by_size or empty_by_rows:
            fed_info["empties"].append(fname)
            file_is_empty.add(fname)

        if stype:
            canonical = normalize_session_type(stype)
            if canonical:
                fed_info["types"].add(canonical)
                fed_info["by_type"].setdefault(canonical, []).append(fname)
                all_types_global.setdefault(canonical, set()).add(fed)

        file_to_type[fname] = normalize_session_type(stype) if stype else None

    key_fed_ids = KEY_INFO.get("fed_ids") or set()
    file_fed_ids = set(per_fed.keys())

    rows = []
    any_issue = False

    # ---- rows for FEDs that actually have files ----
    for fed in sorted(per_fed.keys()):
        info = per_fed[fed]
        file_count = len(info["files"])
        present_types = sorted(info["types"])
        missing_types = sorted(REQUIRED_TYPES - info["types"])
        duplicates = {t: files for t, files in info["by_type"].items() if len(files) > 1}
        count_ok = file_count == 4
        types_ok = len(missing_types) == 0
        has_empties = len(info["empties"]) > 0
        overall_ok = count_ok and types_ok and not has_empties
        if not overall_ok or duplicates:
            any_issue = True

        rows.append(
            {
                "FED": f"FED{fed}",
                "File_Count": file_count,
                "Count_OK(=4)": count_ok,
                "Present_Types": ";".join(present_types),
                "Missing_Types": ";".join(missing_types),
                "Duplicate_Types": "; ".join(
                    f"{t} x{len(files)}" for t, files in info["by_type"].items()
                )
                if duplicates
                else "",
                "Empty_File_Count": len(info["empties"]),
                "Empty_File_Names": "; ".join(sorted(info["empties"]))
                if info["empties"]
                else "",
                "All_Types_OK": types_ok,
                "Overall_OK": overall_ok,
                # key-based info:
                "In_Key": fed in key_fed_ids,
                "In_Files": True,
                "Has_Bandit100": "Bandit100" in info["types"],
                "Has_Bandit80": "Bandit80" in info["types"],
                "Has_FR1": "FR1" in info["types"],
                "Has_PR1": "PR1" in info["types"],
            }
        )

    # ---- rows for FEDs that are in key but have NO files ----
    missing_feds = key_fed_ids - file_fed_ids
    for fed in sorted(missing_feds):
        rows.append(
            {
                "FED": f"FED{fed}",
                "File_Count": 0,
                "Count_OK(=4)": False,
                "Present_Types": "",
                "Missing_Types": ";".join(sorted(REQUIRED_TYPES)),
                "Duplicate_Types": "",
                "Empty_File_Count": 0,
                "Empty_File_Names": "",
                "All_Types_OK": False,
                "Overall_OK": False,
                "In_Key": True,
                "In_Files": False,
                "Has_Bandit100": False,
                "Has_Bandit80": False,
                "Has_FR1": False,
                "Has_PR1": False,
            }
        )

    report = pd.DataFrame(rows)

    unique_devices_total = len(per_fed)
    per_task_counts = {task: len(feds) for task, feds in sorted(all_types_global.items())}

    # ---- summary rows (leave key-based columns blank) ----
    summary_rows = [
        {
            "FED": "SUMMARY",
            "File_Count": unique_devices_total,
            "Count_OK(=4)": "",
            "Present_Types": "",
            "Missing_Types": "",
            "Duplicate_Types": "",
            "Empty_File_Count": "",
            "Empty_File_Names": "",
            "All_Types_OK": "",
            "Overall_OK": "",
            "In_Key": "",
            "In_Files": "",
            "Has_Bandit100": "",
            "Has_Bandit80": "",
            "Has_FR1": "",
            "Has_PR1": "",
        }
    ]
    for task, count in per_task_counts.items():
        summary_rows.append(
            {
                "FED": f"→ {task}",
                "File_Count": count,
                "Count_OK(=4)": "",
                "Present_Types": "",
                "Missing_Types": "",
                "Duplicate_Types": "",
                "Empty_File_Count": "",
                "Empty_File_Names": "",
                "All_Types_OK": "",
                "Overall_OK": "",
                "In_Key": "",
                "In_Files": "",
                "Has_Bandit100": "",
                "Has_Bandit80": "",
                "Has_FR1": "",
                "Has_PR1": "",
            }
        )
    report = pd.concat([report, pd.DataFrame(summary_rows)], ignore_index=True)

    out_xlsx = os.path.join(folder, "fed_session_check_report.xlsx")
    report.to_excel(out_xlsx, index=False)
    style_fed_report_excel(out_xlsx)

    # ---- FED key vs files summary ----
    print("\n=== FED Session Check (Key vs Files) ===")
    print(f"Report saved to: {out_xlsx}")
    print(f"Unique FED IDs in key:   {len(key_fed_ids)}")
    print(f"Unique FED IDs in files: {len(file_fed_ids)}")
    overlap = len(key_fed_ids & file_fed_ids)
    print(f"Overlap (in both):       {overlap}")

    moved_counts = {d: 0 for d in TASK_DIRNAME.values()}
    skipped_unclassified = []
    skipped_empty = []
    for dirname in TASK_DIRNAME.values():
        ensure_dir(os.path.join(folder, dirname))

    for fname in csv_files:
        if fname in file_is_empty:
            skipped_empty.append(fname)
            continue
        stype = file_to_type.get(fname)
        if stype in TASK_DIRNAME:
            if move_file_to_task_folder(folder, fname, stype):
                moved_counts[TASK_DIRNAME[stype]] += 1
        else:
            skipped_unclassified.append(fname)

    print("\n=== FED Session Completeness Report ===")
    print(f"Folder: {folder}")
    print(f"Required session types: {sorted(REQUIRED_TYPES)}")
    print(f"Report saved to: {out_xlsx}\n")
    print(f"Total unique devices (in files): {unique_devices_total}")
    for task, count in per_task_counts.items():
        print(f"  {task}: {count} devices")
    print("\nFile sorting:")
    for dirname, n in moved_counts.items():
        print(f"  Moved into {dirname}: {n} files")
    if skipped_empty:
        print(f"  Skipped empty files (left in place): {len(skipped_empty)}")
    if skipped_unclassified:
        print(
            f"  Skipped unclassified files (no recognized Session_type): {len(skipped_unclassified)}"
        )

    if any_issue:
        print("\nIssues detected:")
        for _, r in report.iterrows():
            if str(r["FED"]).startswith("FED") and (str(r["Overall_OK"]).lower() == "false"):
                print(f"- {r['FED']}:")
                if str(r["Count_OK(=4)"]).lower() == "false":
                    print(f"  * Has {r['File_Count']} files (expected 4).")
                if r["Missing_Types"]:
                    print(f"  * Missing types: {r['Missing_Types']}")
                if r["Duplicate_Types"]:
                    print(f"  * Duplicates: {r['Duplicate_Types']}")
                if str(r["Empty_File_Count"]).strip() not in ("", "0"):
                    print(
                        f"  * Likely empty files ({r['Empty_File_Count']}): {r['Empty_File_Names']}"
                    )
        print("\nOpen the Excel report for the full table (red cells = issues).")
    else:
        print("All FEDs complete, non-empty, and contain all required session types. ✅")

    # ---- Create ZIPs for each FED task folder: Gene_Gene_ID_Task_L0.zip ----
    if KEY_INFO["gene"] and KEY_INFO["gene_id"]:
        gene = KEY_INFO["gene"]
        gene_id = KEY_INFO["gene_id"]
        print("\nCreating FED ZIPs per task folder...")
        for task, dirname in TASK_DIRNAME.items():
            task_dir_path = os.path.join(folder, dirname)
            if not os.path.isdir(task_dir_path):
                continue
            files = [
                f for f in os.listdir(task_dir_path)
                if os.path.isfile(os.path.join(task_dir_path, f))
            ]
            if not files:
                continue
            zip_name = f"{gene}_{gene_id}_{task}_L0.zip"
            zip_path = os.path.join(folder, zip_name)
            try:
                with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                    for fname in files:
                        fpath = os.path.join(task_dir_path, fname)
                        zf.write(fpath, arcname=fname)
                print(f"  Created FED ZIP for {task}: {zip_path}")
            except Exception as e:
                print(f"  ! Failed to create FED zip for {task}: {e}")
    else:
        print("\nNo key loaded; skipping FED ZIP creation (Gene_Gene_ID_Task_L0).")

    # ---- Open folder for user convenience ----
    print("Opening FED folder in file explorer...")
    open_folder_in_explorer(folder)


def run_fed_pipeline():
    """Console wrapper for FED mode."""
    folder = pick_folder()
    return run_fed_pipeline_core(folder)


# =========================
# BEAM implementation
# =========================

BEAM_NAME_RE = re.compile(r"(?i)\bBEAM(\d{3})_(\d{8})(\d{2})\b")


def _print_tree(root, max_entries=60):
    """Print a shallow tree of extracted contents for quick debugging."""
    print(f"\n[Preview of extracted contents under: {root}]")
    shown = 0
    for dirpath, dirnames, filenames in os.walk(root):
        rel = os.path.relpath(dirpath, root)
        indent = "  " * (0 if rel == "." else rel.count(os.sep))
        node_name = os.path.basename(dirpath) if rel != "." else os.path.basename(root)
        print(f"{indent}{node_name}")
        for d in sorted(dirnames):
            print(f"{indent}  /{d}")
            shown += 1
            if shown >= max_entries:
                print("  ... (truncated)")
                return
        for f in sorted(filenames):
            print(f"{indent}   - {f}")
            shown += 1
            if shown >= max_entries:
                print("  ... (truncated)")
                return


def run_beam_pipeline_core(zip_path, start_date, end_date):
    """
    Core BEAM pipeline given zip_path and [start_date, end_date] (YYYYMMDD).
    Also creates beam_session_check_report.xlsx using BEAM IDs in files vs BEAM column in key.
    """
    parent_folder = os.path.dirname(zip_path)
    zip_stem = os.path.splitext(os.path.basename(zip_path))[0]
    extract_root = os.path.join(parent_folder, f"_EXTRACT_{zip_stem}")
    ensure_dir(extract_root)

    print(f"\nSelected ZIP: {zip_path}")
    print(f"Date range to KEEP: {start_date} → {end_date} (inclusive)")

    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            names = zf.namelist()
            print("\nZIP contains (first 30 entries):")
            for n in names[:30]:
                print(f"  - {n}")
            if len(names) > 30:
                print("  ... (truncated)")
            zf.extractall(extract_root)
        print(f"\nExtracted to: {extract_root}")
    except Exception as e:
        print(f"Failed to extract zip: {e}")
        return

    _print_tree(extract_root, max_entries=60)

    def find_csvs(root):
        csv_paths = []
        for r, _, files in os.walk(root):
            for fn in files:
                if fn.lower().endswith(".csv"):
                    csv_paths.append(os.path.join(r, fn))
        return csv_paths

    csv_paths = find_csvs(extract_root)

    nested_dirs_created = []
    if not csv_paths:
        nested_zips = []
        for r, _, files in os.walk(extract_root):
            for fn in files:
                if fn.lower().endswith(".zip"):
                    nested_zips.append(os.path.join(r, fn))
        if nested_zips:
            print(
                f"\nNo CSVs found initially. Found {len(nested_zips)} nested ZIP(s); extracting those..."
            )
            for nz in nested_zips:
                nz_stem = os.path.splitext(os.path.basename(nz))[0]
                nested_out = os.path.join(parent_folder, f"_NESTED_{nz_stem}")
                ensure_dir(nested_out)
                nested_dirs_created.append(nested_out)
                try:
                    with zipfile.ZipFile(nz, "r") as zf:
                        zf.extractall(nested_out)
                    print(f"  Extracted nested zip: {nz} → {nested_out}")
                except Exception as e:
                    print(f"  ! Failed to extract nested zip {nz}: {e}")

            csv_paths = find_csvs(extract_root)
            for d in nested_dirs_created:
                csv_paths.extend(find_csvs(d))

    def in_range(path):
        name_no_ext = os.path.splitext(os.path.basename(path))[0]
        m = BEAM_NAME_RE.search(name_no_ext)
        if not m:
            return False
        date_str = m.group(2)
        return start_date <= date_str <= end_date

    matching_csvs = [p for p in csv_paths if in_range(p)]
    nonmatching = len(csv_paths) - len(matching_csvs)

    if not matching_csvs:
        print("\nNo CSV files matched BEAMXXX_YYYYMMDD?? within your date range.")
        print("Temporary extract folders will be removed.")
        try:
            shutil.rmtree(extract_root, ignore_errors=True)
        except Exception as e:
            print(f"  ! Failed to delete extract root {extract_root}: {e}")
        for d in nested_dirs_created:
            try:
                shutil.rmtree(d, ignore_errors=True)
            except Exception as e:
                print(f"  ! Failed to delete nested folder {d}: {e}")
        return

    # BEAM session check
    file_beam_ids = set()
    for src in matching_csvs:
        m = BEAM_NAME_RE.search(os.path.basename(src))
        if m:
            file_beam_ids.add(m.group(1).zfill(3))

    key_beam_ids = KEY_INFO.get("beam_ids") or set()
    all_beam_ids = sorted(key_beam_ids | file_beam_ids)

    if all_beam_ids:
        rows = []
        for bid in all_beam_ids:
            rows.append(
                {
                    "BEAM_ID": bid,
                    "In_Key": bid in key_beam_ids,
                    "In_Files": bid in file_beam_ids,
                }
            )
        beam_report_df = pd.DataFrame(rows)
        beam_report_path = os.path.join(parent_folder, "beam_session_check_report.xlsx")
        beam_report_df.to_excel(beam_report_path, index=False)
        style_beam_report_excel(beam_report_path)

        print("\n=== BEAM Session Check (Key vs Files) ===")
        print(f"Report saved to: {beam_report_path}")
        print(f"Unique BEAM IDs in key:   {len(key_beam_ids)}")
        print(f"Unique BEAM IDs in files: {len(file_beam_ids)}")
        overlap = len(key_beam_ids & file_beam_ids)
        print(f"Overlap (in both):        {overlap}")
        if not key_beam_ids:
            print("Note: Key did not provide any BEAM IDs (no BEAM column or empty).")
    else:
        print("\nBEAM session check: no BEAM IDs found in key or files; no report created.")

    dest_beam = ensure_dir(os.path.join(parent_folder, "BEAM"))

    moved = 0
    skipped = 0
    for src in matching_csvs:
        try:
            dest = unique_dest_path(dest_beam, os.path.basename(src))
            shutil.move(src, dest)
            moved += 1
        except Exception as e:
            print(f"  ! Could not move {src} → {dest_beam}: {e}")
            skipped += 1

    print("\n=== BEAM Flattening + Date-Range Filter Summary ===")
    print(f"Date range: {start_date} → {end_date}")
    print(f"All CSVs discovered: {len(csv_paths)}")
    print(f"Matched CSVs (kept/moved): {moved}")
    if skipped:
        print(f"  Skipped (errors during move): {skipped}")
    print(f"Non-matching CSVs (discarded when temp folders are deleted): {nonmatching}")
    print(f"\nMatched CSVs are in BEAM folder: {dest_beam}")

    if KEY_INFO["gene"] and KEY_INFO["gene_id"]:
        zip_basename = f"{KEY_INFO['gene']}_{KEY_INFO['gene_id']}_BEAM_L0.zip"
        out_zip_path = os.path.join(parent_folder, zip_basename)
        try:
            with zipfile.ZipFile(out_zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for fname in os.listdir(dest_beam):
                    fpath = os.path.join(dest_beam, fname)
                    if os.path.isfile(fpath):
                        zf.write(fpath, arcname=fname)
            print(f"Created BEAM ZIP: {out_zip_path}")
        except Exception as e:
            print(f"Failed to create BEAM zip {zip_basename}: {e}")
    else:
        print("No key loaded; skipping BEAM ZIP creation (Gene_Gene_ID_BEAM_L0).")

    try:
        shutil.rmtree(extract_root, ignore_errors=True)
    except Exception as e:
        print(f"  ! Failed to delete extract root {extract_root}: {e}")
    for d in nested_dirs_created:
        try:
            shutil.rmtree(d, ignore_errors=True)
        except Exception as e:
            print(f"  ! Failed to delete nested folder {d}: {e}")

    print("Temporary extraction folders removed.")
    print("Done.")

    # ---- Open parent folder so user can see BEAM outputs quickly ----
    print("Opening BEAM parent folder in file explorer...")
    open_folder_in_explorer(parent_folder)


def run_beam_pipeline():
    zip_path = pick_zip_file()
    start_date, end_date = ask_date_range_to_keep()
    return run_beam_pipeline_core(zip_path, start_date, end_date)


# =========================
# GUI application
# =========================

def launch_gui_app():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, simpledialog
    from tkinter import scrolledtext
    from threading import Thread

    root = tk.Tk()
    root.title(f"FilepulldownQC (FED / BEAM)  v{APP_VERSION}")
    root.geometry("900x600")

    class TextRedirector:
        def __init__(self, widget):
            self.widget = widget
            self._orig = getattr(sys, "__stdout__", None)

        def write(self, s):
            try:
                if self.widget is not None:
                    self.widget.insert(tk.END, s)
                    self.widget.see(tk.END)
            except Exception:
                pass
            if self._orig is not None:
                try:
                    self._orig.write(s)
                except Exception:
                    pass

        def flush(self):
            if self._orig is not None:
                try:
                    self._orig.flush()
                except Exception:
                    pass

    top_frame = tk.Frame(root)
    top_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

    key_label_var = tk.StringVar(value="Key: none loaded")
    tk.Label(top_frame, textvariable=key_label_var).pack(side=tk.LEFT, padx=(0, 10))

    def upload_key():
        """Upload Excel key, extract Gene, padded Gene_ID, BEAM IDs, and FED3 IDs."""
        path = filedialog.askopenfilename(
            parent=root,
            title="Select key Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            df = pd.read_excel(path)
        except Exception as e:
            messagebox.showerror("Key load error", f"Could not read key Excel file:\n{e}")
            return
        if df.empty:
            messagebox.showerror("Key load error", "The selected key Excel file is empty.")
            return

        df = df.astype("string")

        col_map = {c.strip().lower(): c for c in df.columns}
        gene_col = col_map.get("gene")
        gene_id_col = col_map.get("gene_id")
        beam_col_name = col_map.get("beam")
        fed_col_name = col_map.get("fed3")

        if not gene_col or not gene_id_col:
            messagebox.showerror(
                "Missing columns",
                "Key Excel file must contain 'Gene' and 'Gene_ID' columns (case-insensitive).",
            )
            return

        def first_nonempty(series):
            for val in series:
                if pd.isna(val):
                    continue
                v = str(val).strip()
                if v:
                    return v
            return None

        gene_val = first_nonempty(df[gene_col])
        gene_id_raw = first_nonempty(df[gene_id_col])
        if not gene_val or not gene_id_raw:
            messagebox.showerror(
                "Empty values",
                "Could not find non-empty values in 'Gene' and 'Gene_ID' columns.",
            )
            return

        gene_id_padded = str(gene_id_raw).strip().zfill(3)

        beam_ids = set()
        if beam_col_name:
            for v in df[beam_col_name]:
                if pd.isna(v):
                    continue
                s = str(v).strip()
                if not s:
                    continue
                m = re.search(r"(\d+)", s)
                if not m:
                    continue
                beam_ids.add(m.group(1).zfill(3))

        fed_ids = set()
        if fed_col_name:
            for v in df[fed_col_name]:
                if pd.isna(v):
                    continue
                s = str(v).strip()
                if not s:
                    continue
                m = re.search(r"(\d+)", s)
                if not m:
                    continue
                fed_ids.add(m.group(1).zfill(3))

        KEY_INFO["gene"] = gene_val
        KEY_INFO["gene_id"] = gene_id_padded
        KEY_INFO["path"] = path
        KEY_INFO["beam_ids"] = beam_ids
        KEY_INFO["fed_ids"] = fed_ids

        gene_str = f"{gene_val}_{gene_id_padded}"
        key_label_var.set(f"Key detected: {gene_str}")
        msg = f"Detected Gene_Gene_ID: {gene_str}"
        msg += f"\nDetected {len(beam_ids)} BEAM IDs in key." if beam_ids else "\nNo BEAM IDs detected in key."
        msg += f"\nDetected {len(fed_ids)} FED IDs in key." if fed_ids else "\nNo FED IDs detected in key."
        messagebox.showinfo("Key loaded", msg)

    tk.Button(top_frame, text="Upload Key (Excel)", command=upload_key).pack(
        side=tk.LEFT, padx=5
    )

    mid_frame = tk.Frame(root)
    mid_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5)

    notebook = ttk.Notebook(mid_frame)
    notebook.pack(fill=tk.BOTH, expand=True)

    fed_tab = tk.Frame(notebook)
    beam_tab = tk.Frame(notebook)
    notebook.add(fed_tab, text="FED")
    notebook.add(beam_tab, text="BEAM")

    # ----- FED controls -----
    tk.Label(
        fed_tab,
        text="Select FED folder containing FED###_*.csv files",
        anchor="w",
    ).pack(anchor="w", padx=5, pady=(5, 2))

    fed_folder_var = tk.StringVar()
    fed_row = tk.Frame(fed_tab)
    fed_row.pack(fill=tk.X, padx=5, pady=(0, 5))
    tk.Entry(fed_row, textvariable=fed_folder_var).pack(side=tk.LEFT, fill=tk.X, expand=True)

    fed_button_row = tk.Frame(fed_tab)
    fed_button_row.pack(anchor="w", padx=5, pady=(0, 5))

    # FED instructions panel
    fed_help = tk.LabelFrame(fed_tab, text="How to run the FED pipeline")
    fed_help.pack(fill=tk.X, padx=5, pady=(0, 8))
    fed_help_text = (
        "1. Upload the Psygene key Excel file first using 'Upload Key (Excel)' at the top "
        "(must include Gene, Gene_ID, FED3, BEAM columns).\n"
        "\n"
        "2. For FED, select the main all_FED_files folder that contains all of the "
        "FED###_*.csv files across Bandit100, Bandit80, FR1, and PR1.\n"
        "\n"
        "3. Click 'Run FED Pipeline'. The app will QC the sessions, create "
        "'fed_session_check_report.xlsx', sort files into task folders, and create "
        "ZIPs: Gene_Gene_ID_Task_L0.zip for each task folder."
    )
    tk.Label(
        fed_help,
        text=fed_help_text,
        anchor="w",
        justify="left",
        wraplength=860,
    ).pack(fill=tk.X, padx=5, pady=3)

    # ----- BEAM controls -----
    tk.Label(beam_tab, text="Select BEAM .zip file", anchor="w").pack(
        anchor="w", padx=5, pady=(5, 2)
    )
    beam_zip_var = tk.StringVar()
    beam_row = tk.Frame(beam_tab)
    beam_row.pack(fill=tk.X, padx=5, pady=(0, 5))
    tk.Entry(beam_row, textvariable=beam_zip_var).pack(
        side=tk.LEFT, fill=tk.X, expand=True
    )

    beam_button_row = tk.Frame(beam_tab)
    beam_button_row.pack(anchor="w", padx=5, pady=(0, 5))

    # BEAM instructions panel
    beam_help = tk.LabelFrame(beam_tab, text="How to run the BEAM pipeline")
    beam_help.pack(fill=tk.X, padx=5, pady=(0, 8))
    beam_help_text = (
        "1. Upload the Psygene key Excel file first using 'Upload Key (Excel)' at the top "
        "(must include Gene, Gene_ID, FED3, BEAM columns).\n"
        "\n"
        "2. For BEAM, select the HubLink export ZIP (or other zip containing the BEAM files). \n"
        "\n"
        "3. Click 'Run BEAM Pipeline'. In the calendar, pick the Saturday the BEAMs "
        "were used (the app will include Saturday and Sunday). \n"
        "The app will filter the files, create a session report, a BEAM folder "
        "with the kept CSVs, and a ZIP named Gene_Gene_ID_BEAM_L0.zip."
    )
    tk.Label(
        beam_help,
        text=beam_help_text,
        anchor="w",
        justify="left",
        wraplength=860,
    ).pack(fill=tk.X, padx=5, pady=3)

    # ----- bottom: progress bar -----
    bottom_container = tk.Frame(root)
    bottom_container.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=5)

    from tkinter import ttk as _ttk
    style = _ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure(
        "green.Horizontal.TProgressbar",
        troughcolor="#ffffff",
        background="#24c140",
        bordercolor="#ffffff",
        lightcolor="#24c140",
        darkcolor="#24c140",
    )

    progress = _ttk.Progressbar(
        bottom_container,
        mode="indeterminate",
        style="green.Horizontal.TProgressbar",
        length=100,
    )
    progress.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=3)

    # ----- log area -----
    log_frame = tk.Frame(root)
    log_frame.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True, padx=5, pady=(5, 0))

    tk.Label(log_frame, text="Log / Summary:").pack(anchor="w")
    log_text = scrolledtext.ScrolledText(log_frame, wrap="word")
    log_text.pack(fill=tk.BOTH, expand=True)

    sys.stdout = TextRedirector(log_text)
    print("FilepulldownQC (FED / BEAM)")

    # ----- GUI-specific pickers -----
    def gui_pick_folder():
        existing = fed_folder_var.get().strip()
        if existing and os.path.isdir(existing):
            return existing
        path = filedialog.askdirectory(
            parent=root, title="Select FED folder containing FED###_*.csv files"
        )
        if not path:
            raise UserCancel("FED folder selection cancelled by user.")
        fed_folder_var.set(path)
        return path

    def gui_pick_zip_file():
        existing = beam_zip_var.get().strip()
        if existing and os.path.isfile(existing) and existing.lower().endswith(".zip"):
            return existing
        path = filedialog.askopenfilename(
            parent=root,
            title="Select BEAM .zip",
            filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")],
        )
        if not path:
            raise UserCancel("BEAM zip selection cancelled by user.")
        if not path.lower().endswith(".zip"):
            messagebox.showerror("Invalid file", "Selected file is not a .zip.")
            raise UserCancel("Invalid BEAM zip; user must re-select.")
        beam_zip_var.set(path)
        return path

    def gui_ask_date_range_to_keep():
        try:
            from tkcalendar import Calendar
        except ImportError:
            pattern = re.compile(r"^\d{8}$")
            while True:
                s = simpledialog.askstring(
                    "Filter by date range",
                    "Enter START date (YYYYMMDD):",
                    parent=root,
                )
                if s is None:
                    raise UserCancel("Date selection cancelled by user.")
                e = simpledialog.askstring(
                    "Filter by date range",
                    "Enter END date (YYYYMMDD):",
                    parent=root,
                )
                if e is None:
                    raise UserCancel("Date selection cancelled by user.")
                s = s.strip()
                e = e.strip()
                if not (pattern.fullmatch(s) and pattern.fullmatch(e)):
                    messagebox.showerror(
                        "Invalid date", "Dates must be exactly 8 digits (YYYYMMDD)."
                    )
                    continue
                if s > e:
                    messagebox.showerror(
                        "Invalid range", "START date must be <= END date."
                    )
                    continue
                return s, e

        from tkcalendar import Calendar

        sel = {"start": None, "end": None}

        top = tk.Toplevel(root)
        top.title("Select BEAM date (Saturday + next day)")
        top.transient(root)
        top.grab_set()

        ttk.Label(
            top,
            text="Select the Saturday the BEAMs were used.\n"
                 "If you pick Saturday, we will include Saturday and Sunday.",
        ).pack(padx=10, pady=(10, 5))

        cal = Calendar(top, selectmode="day")
        cal.pack(padx=10, pady=5)

        def on_ok():
            d = cal.selection_get()
            if d is None:
                messagebox.showerror("No date", "Please select a date.")
                return
            start = d
            if d.weekday() == 5:
                end = d + timedelta(days=1)
            else:
                end = d
            sel["start"] = start.strftime("%Y%m%d")
            sel["end"] = end.strftime("%Y%m%d")
            top.destroy()

        def on_cancel():
            sel["start"] = None
            sel["end"] = None
            top.destroy()

        btn_frame = tk.Frame(top)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="OK", command=on_ok).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=on_cancel).pack(
            side=tk.LEFT, padx=5
        )

        root.wait_window(top)
        if not sel["start"]:
            raise UserCancel("Date selection cancelled by user.")
        return sel["start"], sel["end"]

    fed_run_btn = None
    beam_run_btn = None

    def set_run_buttons_state(state):
        if fed_run_btn is not None:
            fed_run_btn.config(state=state)
        if beam_run_btn is not None:
            beam_run_btn.config(state=state)

    def monitor_thread(thread):
        if thread.is_alive():
            root.after(100, lambda: monitor_thread(thread))
        else:
            progress.stop()
            set_run_buttons_state("normal")
            root.update_idletasks()

    def run_fed_clicked():
        try:
            folder = gui_pick_folder()
        except UserCancel as uc:
            print(str(uc))
            return

        print("\n===== Running FED pipeline =====")
        set_run_buttons_state("disabled")
        progress.start(10)
        root.update_idletasks()

        def worker():
            try:
                run_fed_pipeline_core(folder)
            except Exception as e:
                print(f"FED pipeline error: {e}")

        t = Thread(target=worker, daemon=True)
        t.start()
        monitor_thread(t)

    def run_beam_clicked():
        try:
            zip_path = gui_pick_zip_file()
            start_date, end_date = gui_ask_date_range_to_keep()
        except UserCancel as uc:
            print(str(uc))
            return

        print("\n===== Running BEAM pipeline =====")
        set_run_buttons_state("disabled")
        progress.start(10)
        root.update_idletasks()

        def worker():
            try:
                run_beam_pipeline_core(zip_path, start_date, end_date)
            except Exception as e:
                print(f"BEAM pipeline error: {e}")

        t = Thread(target=worker, daemon=True)
        t.start()
        monitor_thread(t)

    tk.Button(fed_button_row, text="Browse...", command=gui_pick_folder).pack(
        side=tk.LEFT, padx=(0, 5)
    )
    fed_run_btn = tk.Button(fed_button_row, text="Run FED Pipeline", command=run_fed_clicked)
    fed_run_btn.pack(side=tk.LEFT)

    tk.Button(beam_button_row, text="Browse...", command=gui_pick_zip_file).pack(
        side=tk.LEFT, padx=(0, 5)
    )
    beam_run_btn = tk.Button(beam_button_row, text="Run BEAM Pipeline", command=run_beam_clicked)
    beam_run_btn.pack(side=tk.LEFT)

    root.mainloop()


# =========================
# Main
# =========================

def main():
    try:
        launch_gui_app()
    except Exception as e:
        if hasattr(sys, "__stdout__") and sys.__stdout__ is not None:
            sys.stdout = sys.__stdout__

        print(f"GUI failed ({e}), falling back to console mode.")
        mode = input("Select mode: [F]ED or [B]EAM? ").strip().lower()
        if mode.startswith("f"):
            run_fed_pipeline()
        elif mode.startswith("b"):
            run_beam_pipeline()
        else:
            print("No valid mode selected. Exiting.")
            sys.exit(0)


if __name__ == "__main__":
    main()
